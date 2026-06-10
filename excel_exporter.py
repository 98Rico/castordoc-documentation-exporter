"""
excel_exporter.py

Generate a dbt-oriented model specification workbook from CastorDoc TXT exports.

Design
------
The exporter has two separate responsibilities:

1. export_castordoc.py
   - extracts CastorDoc pages broadly;
   - follows useful ReadMe links, including [C], [A], [AC], related facts, etc.;
   - writes TXT evidence files.

2. excel_exporter.py
   - interprets those TXT files;
   - builds a dbt model specification workbook;
   - only treats [A] and [AC] items as target model columns;
   - keeps [C], fact pages, and related attributes as evidence/dependencies.

Important column rule
---------------------
Only [A] and [AC] items are target column candidates.

Example:
    ✅ [C] Production Work In Progress Step   -> concept / dependency, not a column
    ✅ [A] Work In Progress Number            -> column
    ✅ [A] Work In Progress Label             -> column
    ✅ [AC] Some Attribute Concept            -> column candidate

For pages where the root model page contains no [A]/[AC], the generator looks at
direct linked pages at depth 1. This handles the common CastorDoc pattern:

    root page = model requirement / test page
    depth 1 page = business concept page containing [A]/[AC] attributes
    depth 2+ pages = fact pages / attribute detail pages / related evidence

Output
------
    castordoc_model_specification.xlsx

Usage
-----
    uv run python excel_exporter.py \
      --export-folder "Output/castordoc_export_YYYY-MM-DD_HH-MM-SS"
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# DATA STRUCTURES
# ============================================================


@dataclass
class DocumentationPage:
    file_name: str
    title: str
    url: str
    source_type: str
    depth: int | None
    parent_url: str
    exported_at: str
    content: str
    summary: str


@dataclass
class ColumnSpec:
    column_name: str
    business_name: str
    suggested_dbt_name: str
    suggested_type: str
    nullable: str
    source_dependency: str
    calculation_rule: str
    business_definition: str
    example_values: str
    data_quality_rules: str
    open_questions: str
    evidence_page: str
    evidence_url: str


@dataclass
class RequirementSpec:
    requirement_id: str
    requirement_type: str
    requirement: str
    expected_result: str
    priority: str
    evidence_page: str
    evidence_url: str


@dataclass
class RuleSpec:
    rule_id: str
    field_or_metric: str
    rule_type: str
    formula_or_logic: str
    dependencies: str
    implementation_notes: str
    evidence_page: str
    evidence_url: str


@dataclass
class TestSpec:
    test_name: str
    columns: str
    test_type: str
    severity: str
    expected_result: str
    suggested_dbt_test: str
    evidence_page: str
    evidence_url: str


@dataclass
class DependencySpec:
    target_model: str
    dependency_type: str
    dependency_name: str
    dependency_role: str
    join_or_usage: str
    evidence_page: str
    evidence_url: str


# ============================================================
# CLI
# ============================================================


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a dbt model specification workbook from CastorDoc TXT exports."
    )

    parser.add_argument(
        "--export-folder",
        required=True,
        help="Folder containing CastorDoc .txt exports and export_summary.txt.",
    )

    parser.add_argument(
        "--output-file",
        default=None,
        help="Optional output xlsx path.",
    )

    return parser.parse_args()


# ============================================================
# TEXT HELPERS
# ============================================================


def clean_excel_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    text = text.replace("\xa0", " ")
    text = text.replace("\r", "")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen = set()
    output = []

    for value in values:
        clean = value.strip()

        if not clean:
            continue

        key = clean.lower()

        if key not in seen:
            seen.add(key)
            output.append(clean)

    return output


def parse_metadata_and_content(text: str) -> tuple[dict[str, str], str]:
    metadata: dict[str, str] = {}
    lines = text.splitlines()
    content_start_index = 0

    for index, line in enumerate(lines):
        if not line.strip():
            content_start_index = index + 1
            break

        if ":" in line:
            key, value = line.split(":", 1)
            metadata[key.strip().upper()] = value.strip()

    content = "\n".join(lines[content_start_index:]).strip()
    return metadata, content


def first_non_empty_lines(text: str, max_lines: int = 4) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    return "\n".join(lines[:max_lines])


def parse_documentation_file(path: Path) -> DocumentationPage:
    raw_text = path.read_text(encoding="utf-8")
    metadata, content = parse_metadata_and_content(raw_text)

    try:
        depth = int(metadata.get("DEPTH", ""))
    except ValueError:
        depth = None

    return DocumentationPage(
        file_name=path.name,
        title=metadata.get("TITLE", path.stem),
        url=metadata.get("URL", ""),
        source_type=metadata.get("SOURCE_TYPE", ""),
        depth=depth,
        parent_url=metadata.get("PARENT_URL", ""),
        exported_at=metadata.get("EXPORTED_AT", ""),
        content=clean_excel_text(content),
        summary=first_non_empty_lines(content),
    )


def parse_summary_file(summary_path: Path) -> list[tuple[str, str]]:
    if not summary_path.exists():
        return []

    rows = []

    for line in summary_path.read_text(encoding="utf-8").splitlines():
        if ":" not in line:
            continue

        key, value = line.split(":", 1)
        rows.append((key.strip(), value.strip()))

    return rows


def get_summary_value(summary_rows: list[tuple[str, str]], key: str, default: str = "") -> str:
    return dict(summary_rows).get(key, default)


# ============================================================
# SECTION EXTRACTION
# ============================================================


def common_stop_patterns() -> list[str]:
    return [
        r"\n\s*🧾?\s*examples?\s+of\s+values",
        r"\n\s*🛠️?\s*key\s+attributes",
        r"\n\s*🔁?\s*update\s+frequency",
        r"\n\s*🔁?\s*lifecycle",
        r"\n\s*📐?\s*business\s+rules",
        r"\n\s*🏛️?\s*golden\s+source",
        r"\n\s*🧭?\s*scope",
        r"\n\s*⚠️?\s*additional\s+notes",
        r"\n\s*pinned\s+assets",
        r"\n\s*functional\s+tests",
        r"\n\s*concepts?\s*:",
    ]


def extract_section(content: str, start_patterns: list[str], stop_patterns: list[str]) -> str:
    if not content:
        return ""

    start_match = None

    for pattern in start_patterns:
        match = re.search(pattern, content, re.IGNORECASE)

        if match and (start_match is None or match.start() < start_match.start()):
            start_match = match

    if not start_match:
        return ""

    body_start = start_match.end()
    stop_index = len(content)

    for pattern in stop_patterns:
        match = re.search(pattern, content[body_start:], re.IGNORECASE)

        if match:
            absolute = body_start + match.start()
            if absolute < stop_index:
                stop_index = absolute

    return clean_excel_text(content[body_start:stop_index])


def extract_definition(content: str) -> str:
    return extract_section(
        content,
        start_patterns=[
            r"definition\s*&\s*purpose",
            r"coalesce\s+catalog\s+description",
        ],
        stop_patterns=common_stop_patterns(),
    )


def extract_examples(content: str) -> str:
    return extract_section(
        content,
        start_patterns=[
            r"examples?\s+of\s+values",
            r"examples?\s+of\s+value",
        ],
        stop_patterns=[
            r"\n\s*🛠️?\s*key\s+attributes",
            r"\n\s*🔁?\s*update\s+frequency",
            r"\n\s*📐?\s*business\s+rules",
            r"\n\s*🏛️?\s*golden\s+source",
            r"\n\s*⚠️?\s*additional\s+notes",
            r"\n\s*pinned\s+assets",
        ],
    )


def extract_business_rules(content: str) -> str:
    return extract_section(
        content,
        start_patterns=[
            r"business\s+rules\s*&\s*data\s+quality",
            r"business\s+rules",
            r"data\s+quality",
        ],
        stop_patterns=[
            r"\n\s*🏛️?\s*golden\s+source",
            r"\n\s*⚠️?\s*additional\s+notes",
            r"\n\s*🧭?\s*scope",
            r"\n\s*pinned\s+assets",
        ],
    )


def extract_golden_source(content: str) -> str:
    return extract_section(
        content,
        start_patterns=[r"golden\s+source"],
        stop_patterns=[
            r"\n\s*⚠️?\s*additional\s+notes",
            r"\n\s*🧭?\s*scope",
            r"\n\s*pinned\s+assets",
        ],
    )


def extract_update_frequency(content: str) -> str:
    return extract_section(
        content,
        start_patterns=[r"update\s+frequency"],
        stop_patterns=[
            r"\n\s*🔁?\s*lifecycle",
            r"\n\s*📐?\s*business\s+rules",
            r"\n\s*🏛️?\s*golden\s+source",
            r"\n\s*⚠️?\s*additional\s+notes",
            r"\n\s*pinned\s+assets",
        ],
    )


# ============================================================
# CASTORDOC ITEM PARSING
# ============================================================


def extract_castordoc_items(content: str, item_types: set[str] | None = None) -> list[str]:
    """
    Extract typed CastorDoc items.

    Examples:
        ✅ [C] Production Work In Progress Step
        ✅ [A] Work In Progress Number
        ✅ [AC] Some Attribute Concept
        ✅ [Dim] Work In Progress Step

    item_types:
        {"A", "AC"} -> target column candidates
        {"C"}       -> concepts
        {"Dim"}     -> dimension model labels
        None        -> all typed items
    """
    results = []

    for line in content.splitlines():
        clean = line.strip()
        match = re.search(r"✅\s*\[([A-Za-z]+)\]\s*(.+)$", clean)

        if not match:
            continue

        detected_type = match.group(1).strip()
        label = match.group(2).strip()

        if item_types is not None and detected_type not in item_types:
            continue

        results.append(label)

    return dedupe_keep_order(results)


def get_root_pages(pages: list[DocumentationPage]) -> list[DocumentationPage]:
    roots = [page for page in pages if page.source_type == "root"]

    if roots:
        return roots

    min_depth = min((page.depth or 0 for page in pages), default=0)
    return [page for page in pages if (page.depth or 0) == min_depth]


def get_direct_children_of_roots(pages: list[DocumentationPage]) -> list[DocumentationPage]:
    roots = get_root_pages(pages)
    root_urls = {page.url for page in roots}

    direct_children = [
        page
        for page in pages
        if page.parent_url in root_urls or page.depth == 1
    ]

    return direct_children


def is_related_fact_page(page: DocumentationPage) -> bool:
    content_lower = page.content.lower()
    url_lower = page.url.lower()

    return (
        "/fact-" in url_lower
        or "production order number" in content_lower
        or "is current step" in content_lower
        or "confirmed production" in content_lower
        or "remaining production" in content_lower
        or "theoritical production" in content_lower
        or "theoretical production" in content_lower
    )


def collect_target_model_columns(pages: list[DocumentationPage]) -> list[str]:
    """
    Collect target model columns.

    Priority:
    1. [A]/[AC] from root page.
    2. [A]/[AC] from direct child pages of root.
    3. Fallback to root example headers.

    Never collect:
    - [C] concepts;
    - fact-page attributes;
    - WIP label values like "Production Order Delivered (18)".
    """
    roots = get_root_pages(pages)

    columns = []

    for root in roots:
        columns.extend(extract_castordoc_items(root.content, item_types={"A", "AC"}))

    if columns:
        return dedupe_keep_order(columns)

    for child in get_direct_children_of_roots(pages):
        if is_related_fact_page(child):
            continue

        columns.extend(extract_castordoc_items(child.content, item_types={"A", "AC"}))

    if columns:
        return dedupe_keep_order(columns)

    # Last fallback: root example headers only.
    fallback_columns = []

    for root in roots:
        examples = extract_examples(root.content)

        for line in examples.splitlines():
            clean = line.strip()

            if looks_like_column_header(clean):
                fallback_columns.append(clean)

    return dedupe_keep_order(fallback_columns)


def looks_like_column_header(line: str) -> bool:
    clean = line.strip()
    lower = clean.lower()

    if not clean:
        return False

    if len(clean) > 90:
        return False

    if re.fullmatch(r"\d+[,.]?\d*", clean):
        return False

    if re.search(r"\(\d+\)$", clean):
        return False

    if lower in {"true", "false", "day", "sap"}:
        return False

    keywords = [
        "number",
        "label",
        "step",
        "date",
        "time",
        "quantity",
        "hours",
        "days",
        "unit",
        "current",
        "start",
        "end",
        "lead",
        "remaining",
        "confirmed",
        "theoretical",
        "theoritical",
    ]

    return any(keyword in lower for keyword in keywords)


# ============================================================
# MODEL INFERENCE
# ============================================================


def to_dbt_column_name(name: str) -> str:
    text = name.strip()
    text = text.replace("Theoritical", "Theoretical")
    text = text.replace("realased", "released")
    text = re.sub(r"[^A-Za-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_").lower()


def infer_target_model_name(pages: list[DocumentationPage]) -> str:
    """
    Infer model name from the root page content.

    Example:
        ✅ [Dim] Work In Progress Step -> dim_work_in_progress_step
    """
    roots = get_root_pages(pages)

    for root in roots:
        dim_items = extract_castordoc_items(root.content, item_types={"Dim", "DIM"})
        fact_items = extract_castordoc_items(root.content, item_types={"Fact", "Fct", "FACT", "FCT"})

        if dim_items:
            return "dim_" + to_dbt_column_name(dim_items[0])

        if fact_items:
            return "fct_" + to_dbt_column_name(fact_items[0])

    return "to_be_confirmed"


def find_best_model_definition_page(pages: list[DocumentationPage]) -> DocumentationPage:
    """
    Prefer a direct child concept page with a clean business definition.

    Root pages often contain acceptance criteria and test definitions, so they are
    not always the best source for business purpose.
    """
    roots = get_root_pages(pages)
    direct_children = get_direct_children_of_roots(pages)

    candidates = []

    for page in direct_children + roots:
        definition = extract_definition(page.content)

        if not definition:
            continue

        score = 0

        if page in direct_children:
            score += 50

        if extract_castordoc_items(page.content, item_types={"A", "AC"}):
            score += 30

        if extract_castordoc_items(page.content, item_types={"C"}):
            score += 20

        if "functional tests" in page.content.lower():
            score -= 30

        if is_related_fact_page(page):
            score -= 40

        score -= min(len(definition) // 1000, 10)

        candidates.append((score, page))

    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        return candidates[0][1]

    return roots[0] if roots else pages[0]


# ============================================================
# COLUMN SPEC INFERENCE
# ============================================================


def infer_type_from_name_and_examples(name: str, examples: str = "") -> str:
    lower_name = name.lower()

    if lower_name.startswith("is_") or lower_name.startswith("has_"):
        return "BOOLEAN"

    if "date" in lower_name or "time" in lower_name or "start" in lower_name or "end" in lower_name:
        return "DATE / TIMESTAMP_NTZ"

    if "quantity" in lower_name or "hours" in lower_name or "days" in lower_name or "lead_time" in lower_name:
        return "NUMBER"

    if "number" in lower_name:
        if "production_order" in lower_name:
            return "VARCHAR"
        return "NUMBER / INTEGER"

    return "VARCHAR"


def find_best_evidence_page(column_name: str, pages: list[DocumentationPage]) -> DocumentationPage | None:
    """
    Find the page that best documents this attribute.

    For current WIP export:
    - Work In Progress Number -> its attribute detail page
    - Work In Progress Label  -> its attribute detail page
    """
    lower_column = column_name.lower()
    candidates = []

    for page in pages:
        content_lower = page.content.lower()

        score = 0

        if f"{lower_column} is " in content_lower:
            score += 100

        if lower_column in content_lower:
            score += 20

        if page.depth == 2:
            score += 10

        if is_related_fact_page(page):
            score -= 40

        if extract_castordoc_items(page.content, item_types={"A", "AC"}):
            score += 5

        if score > 0:
            candidates.append((score, page))

    if candidates:
        candidates.sort(key=lambda item: item[0], reverse=True)
        return candidates[0][1]

    roots = get_root_pages(pages)
    return roots[0] if roots else None


def extract_example_values_for_column(evidence_page: DocumentationPage | None) -> str:
    if not evidence_page:
        return "Not documented in CastorDoc export"

    examples = extract_examples(evidence_page.content)

    if not examples:
        return "Not documented in CastorDoc export"

    lines = [line.strip() for line in examples.splitlines() if line.strip()]
    return "\n".join(lines[:30]) if lines else "Not documented in CastorDoc export"


def build_business_definition(evidence_page: DocumentationPage | None) -> str:
    if not evidence_page:
        return "Not documented in CastorDoc export"

    definition = extract_definition(evidence_page.content)
    return definition or evidence_page.summary or "Not documented in CastorDoc export"


def infer_source_dependency(evidence_page: DocumentationPage | None) -> str:
    if not evidence_page:
        return "Not documented in CastorDoc export"

    golden_source = extract_golden_source(evidence_page.content)

    if golden_source:
        return golden_source

    lower = evidence_page.content.lower()

    if "sap" in lower:
        return "SAP"

    if "manual" in lower or "entered manually" in lower:
        return "Data to be entered manually"

    return "Not documented in CastorDoc export"


def build_data_quality_rules(column_name: str, evidence_page: DocumentationPage | None) -> str:
    documented = extract_business_rules(evidence_page.content) if evidence_page else ""
    inferred = []
    lower = column_name.lower()

    if "work in progress number" in lower:
        inferred.append("Expected official WIP step numbers are 1 to 18.")

    if "number" in lower:
        inferred.append("Should be not null where required by the business grain.")

    if "label" in lower:
        inferred.append("Should be not null and human-readable.")

    combined = []

    if documented:
        combined.append(documented)

    combined.extend(inferred)

    return "\n".join(dedupe_keep_order(combined)) or "Not documented in CastorDoc export"


def infer_nullable(column_name: str, evidence_page: DocumentationPage | None) -> str:
    lower = column_name.lower()
    content = evidence_page.content.lower() if evidence_page else ""

    if "cannot be null" in content or "mandatory" in content:
        return "No"

    if "number" in lower or "label" in lower:
        return "No / to confirm"

    return "To be confirmed"


def infer_calculation_rule(column_name: str) -> str:
    lower = column_name.lower()

    if "number" in lower or "label" in lower:
        return "No calculation documented; reference attribute / manual dimension value."

    return "Not documented in CastorDoc export."


def build_open_questions(column_name: str, calculation_rule: str, evidence_page: DocumentationPage | None) -> str:
    questions = []

    if not evidence_page:
        questions.append("Confirm business definition and source.")

    if "to confirm" in calculation_rule.lower():
        questions.append(f"Confirm derivation rule for {column_name}.")

    return "\n".join(dedupe_keep_order(questions))


def build_column_specs(pages: list[DocumentationPage]) -> list[ColumnSpec]:
    column_names = collect_target_model_columns(pages)
    specs = []

    for column_name in column_names:
        evidence_page = find_best_evidence_page(column_name, pages)
        examples = extract_example_values_for_column(evidence_page)
        suggested_dbt_name = to_dbt_column_name(column_name)
        calculation_rule = infer_calculation_rule(column_name)

        specs.append(
            ColumnSpec(
                column_name=column_name,
                business_name=column_name,
                suggested_dbt_name=suggested_dbt_name,
                suggested_type=infer_type_from_name_and_examples(suggested_dbt_name, examples),
                nullable=infer_nullable(column_name, evidence_page),
                source_dependency=infer_source_dependency(evidence_page),
                calculation_rule=calculation_rule,
                business_definition=build_business_definition(evidence_page),
                example_values=examples,
                data_quality_rules=build_data_quality_rules(column_name, evidence_page),
                open_questions=build_open_questions(column_name, calculation_rule, evidence_page),
                evidence_page=display_page_name(evidence_page) if evidence_page else "",
                evidence_url=evidence_page.url if evidence_page else "",
            )
        )

    return specs


# ============================================================
# REQUIREMENTS / RULES / TESTS / DEPENDENCIES
# ============================================================


def build_requirement_specs(pages: list[DocumentationPage]) -> list[RequirementSpec]:
    requirements = []
    counter = 1

    for page in pages:
        if not re.search(r"acceptance\s+criteria|functional\s+tests", page.content, re.IGNORECASE):
            continue

        lines = [line.strip() for line in page.content.splitlines() if line.strip()]
        cleaned = []
        in_functional_area = False

        skip_values = {
            "functional tests definition",
            "description",
            "acceptance criteria",
            "expected result",
        }

        for line in lines:
            lower = line.lower()

            if "functional tests" in lower:
                in_functional_area = True
                continue

            if not in_functional_area:
                continue

            if lower in skip_values:
                continue

            if lower.startswith("other acceptance criteria"):
                continue

            if lower.startswith("pinned assets"):
                break

            cleaned.append(line)

        for i in range(0, len(cleaned), 3):
            chunk = cleaned[i:i + 3]

            if len(chunk) == 3:
                requirements.append(
                    RequirementSpec(
                        requirement_id=f"REQ-{counter:03d}",
                        requirement_type="Functional / Acceptance Criteria",
                        requirement=chunk[1],
                        expected_result=chunk[2],
                        priority="High",
                        evidence_page=display_page_name(page),
                        evidence_url=page.url,
                    )
                )
                counter += 1

        for line in lines:
            if line.lower().startswith("validate that"):
                requirements.append(
                    RequirementSpec(
                        requirement_id=f"REQ-{counter:03d}",
                        requirement_type="Additional Acceptance Criteria",
                        requirement=line,
                        expected_result="To be validated in dbt/Sifflet.",
                        priority="Medium",
                        evidence_page=display_page_name(page),
                        evidence_url=page.url,
                    )
                )
                counter += 1

    return requirements


def build_rule_specs(column_specs: list[ColumnSpec]) -> list[RuleSpec]:
    rules = []
    counter = 1

    for col in column_specs:
        rules.append(
            RuleSpec(
                rule_id=f"RULE-{counter:03d}",
                field_or_metric=col.suggested_dbt_name,
                rule_type="Mapping / Reference",
                formula_or_logic=col.calculation_rule,
                dependencies=col.source_dependency,
                implementation_notes="Implement as documented reference/dimension attribute.",
                evidence_page=col.evidence_page,
                evidence_url=col.evidence_url,
            )
        )
        counter += 1

        if col.data_quality_rules:
            rules.append(
                RuleSpec(
                    rule_id=f"RULE-{counter:03d}",
                    field_or_metric=col.suggested_dbt_name,
                    rule_type="Data Quality Rule",
                    formula_or_logic=col.data_quality_rules,
                    dependencies=col.source_dependency,
                    implementation_notes="Translate into dbt tests where possible.",
                    evidence_page=col.evidence_page,
                    evidence_url=col.evidence_url,
                )
            )
            counter += 1

    return rules


def build_test_specs(column_specs: list[ColumnSpec], requirements: list[RequirementSpec]) -> list[TestSpec]:
    tests = []

    column_names = {col.suggested_dbt_name for col in column_specs}

    for col in column_specs:
        dbt_name = col.suggested_dbt_name

        if col.nullable.lower().startswith("no"):
            tests.append(
                TestSpec(
                    test_name=f"not_null_{dbt_name}",
                    columns=dbt_name,
                    test_type="not_null",
                    severity="error",
                    expected_result=f"{dbt_name} should not be null.",
                    suggested_dbt_test="not_null",
                    evidence_page=col.evidence_page,
                    evidence_url=col.evidence_url,
                )
            )

        if dbt_name == "work_in_progress_number":
            tests.append(
                TestSpec(
                    test_name="accepted_values_work_in_progress_number",
                    columns=dbt_name,
                    test_type="accepted_values",
                    severity="error",
                    expected_result="Only official WIP step numbers 1 to 18 are expected.",
                    suggested_dbt_test="accepted_values: [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18]",
                    evidence_page=col.evidence_page,
                    evidence_url=col.evidence_url,
                )
            )

        if dbt_name == "work_in_progress_label":
            tests.append(
                TestSpec(
                    test_name="not_empty_work_in_progress_label",
                    columns=dbt_name,
                    test_type="expression_is_true",
                    severity="error",
                    expected_result="work_in_progress_label should be human-readable and not empty.",
                    suggested_dbt_test="dbt_utils.expression_is_true: LENGTH(TRIM(work_in_progress_label)) > 0",
                    evidence_page=col.evidence_page,
                    evidence_url=col.evidence_url,
                )
            )

    for req in requirements:
        lower_req = req.requirement.lower()

        if "18" in lower_req and "work in progress" in lower_req:
            tests.append(
                TestSpec(
                    test_name="expected_18_wip_steps",
                    columns="work_in_progress_number",
                    test_type="cardinality / accepted_values",
                    severity="error",
                    expected_result="The WIP dimension should contain exactly the official 18 steps.",
                    suggested_dbt_test="accepted_values + row count / custom singular test",
                    evidence_page=req.evidence_page,
                    evidence_url=req.evidence_url,
                )
            )

        if "unique" in lower_req and "work in progress" in lower_req:
            tests.append(
                TestSpec(
                    test_name="unique_work_in_progress_number_label",
                    columns="work_in_progress_number, work_in_progress_label",
                    test_type="unique_combination",
                    severity="error",
                    expected_result="No duplicate work_in_progress_number + work_in_progress_label.",
                    suggested_dbt_test="dbt_utils.unique_combination_of_columns",
                    evidence_page=req.evidence_page,
                    evidence_url=req.evidence_url,
                )
            )

        if "referential integrity" in lower_req or "referenced in gold fact tables" in lower_req:
            tests.append(
                TestSpec(
                    test_name="relationships_wip_step_with_fact_tables",
                    columns="work_in_progress_number / work_in_progress_label",
                    test_type="relationships",
                    severity="error",
                    expected_result="No orphan WIP step references exist in fact tables.",
                    suggested_dbt_test="relationships or custom singular test against fact tables",
                    evidence_page=req.evidence_page,
                    evidence_url=req.evidence_url,
                )
            )

    # Dedupe by test name.
    deduped = {}

    for test in tests:
        deduped[test.test_name] = test

    return list(deduped.values())


def classify_dependency(page: DocumentationPage, target_column_names: set[str]) -> tuple[str, str]:
    content_lower = page.content.lower()

    attributes = extract_castordoc_items(page.content, item_types={"A", "AC"})
    concepts = extract_castordoc_items(page.content, item_types={"C"})

    if page.source_type == "root":
        return "root_model_page", "Root model documentation and acceptance criteria."

    if is_related_fact_page(page):
        return "related_fact_or_attribute", "Related fact or fact-level attribute evidence; not a target dimension column."

    matching_target_attrs = [
        attr for attr in attributes if to_dbt_column_name(attr) in target_column_names
    ]

    if matching_target_attrs:
        return "target_attribute_context", "Documents target model attributes: " + ", ".join(matching_target_attrs)

    if concepts:
        return "business_concept", "Business concept evidence: " + ", ".join(concepts)

    if "work in progress number is" in content_lower or "work in progress label is" in content_lower:
        return "target_attribute_detail", "Detailed definition/rules for a target model attribute."

    return "linked_documentation", "Linked documentation evidence."


def build_dependency_specs(
    pages: list[DocumentationPage],
    target_model: str,
    column_specs: list[ColumnSpec],
) -> list[DependencySpec]:
    dependencies = []
    target_column_names = {col.suggested_dbt_name for col in column_specs}

    for page in pages:
        dependency_type, usage = classify_dependency(page, target_column_names)

        definition = extract_definition(page.content)
        concepts = extract_castordoc_items(page.content, item_types={"C"})
        attributes = extract_castordoc_items(page.content, item_types={"A", "AC"})
        golden_source = extract_golden_source(page.content)

        role_parts = []

        if definition:
            role_parts.append(definition)

        if concepts:
            role_parts.append("Concepts: " + ", ".join(concepts))

        if attributes:
            role_parts.append("Attributes: " + ", ".join(attributes))

        if golden_source:
            role_parts.append("Golden source: " + golden_source)

        dependencies.append(
            DependencySpec(
                target_model=target_model,
                dependency_type=dependency_type,
                dependency_name=display_page_name(page),
                dependency_role="\n".join(role_parts) or page.summary,
                join_or_usage=usage,
                evidence_page=display_page_name(page),
                evidence_url=page.url,
            )
        )

    return dependencies


def build_implementation_notes(
    target_model: str,
    column_specs: list[ColumnSpec],
    requirements: list[RequirementSpec],
    tests: list[TestSpec],
) -> list[tuple[str, str]]:
    column_names = [col.suggested_dbt_name for col in column_specs]

    if "work_in_progress_number" in column_names and "work_in_progress_label" in column_names:
        grain = "One row per work_in_progress_number + work_in_progress_label."
        primary_key = "work_in_progress_number + work_in_progress_label"
    else:
        grain = "To be confirmed from business documentation."
        primary_key = "To be confirmed."

    return [
        ("Target model", target_model),
        ("Recommended dbt layer", "03_marts if BI-ready; 02_intermediate if source-truth/rules still need validation."),
        ("Suggested grain", grain),
        ("Suggested primary key", primary_key),
        ("Column extraction rule", "Only [A] and [AC] items from target model context are treated as columns. [C] pages remain dependencies/evidence."),
        ("Number of documented columns", str(len(column_specs))),
        ("Number of extracted requirements", str(len(requirements))),
        ("Number of suggested data tests", str(len(tests))),
        ("Implementation guidance", "Implement only documented or business-confirmed rules. Keep undocumented formulas as TODO/open questions."),
    ]


# ============================================================
# DISPLAY NAME HELPERS
# ============================================================


def display_page_name(page: DocumentationPage | None) -> str:
    if not page:
        return ""

    if page.title and page.title.lower() != "details":
        return page.title

    dim_items = extract_castordoc_items(page.content, item_types={"Dim", "DIM"})
    if dim_items:
        return "[Dim] " + dim_items[0]

    fact_items = extract_castordoc_items(page.content, item_types={"Fact", "Fct", "FACT", "FCT"})
    if fact_items:
        return "[Fact] " + fact_items[0]

    attr_items = extract_castordoc_items(page.content, item_types={"A", "AC"})
    if attr_items:
        return "[A] " + ", ".join(attr_items)

    concept_items = extract_castordoc_items(page.content, item_types={"C"})
    if concept_items:
        return "[C] " + ", ".join(concept_items)

    first_line = first_non_empty_lines(page.content, 1)
    if first_line and first_line.lower() != "coalesce catalog description":
        return first_line[:80]

    return page.file_name


# ============================================================
# EXCEL FORMATTING
# ============================================================

HEADER_FILL = "1F4E78"
WHITE_FONT = "FFFFFF"
BORDER_COLOR = "D9E2F3"

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def write_header_row(ws, headers: list[str], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=header)


def style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color=WHITE_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    if max_row < min_row:
        return

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: dict[str, int]) -> None:
    for column_letter, width in widths.items():
        ws.column_dimensions[column_letter].width = width


def add_table(ws, table_name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> None:
    if end_row <= start_row:
        return

    ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def finalize_sheet(ws, freeze_cell: str = "A2") -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze_cell

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 36


def build_table_sheet(
    wb: Workbook,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: dict[str, int],
    table_name: str,
) -> None:
    """
    Build one Excel sheet.

    Important:
    - Row 1 is always the header.
    - Data starts at row 2.
    - This prevents Excel table repair issues.
    """
    ws = wb.create_sheet(sheet_name)

    write_header_row(ws, headers, row=1)
    style_header_row(ws, 1, 1, len(headers))

    if rows:
        for row in rows:
            ws.append(row)
    else:
        ws.append(["No data"] + [""] * (len(headers) - 1))

    end_row = 1 + max(len(rows), 1)

    add_table(ws, table_name, 1, 1, end_row, len(headers))
    style_range(ws, 2, end_row, 1, len(headers))
    set_widths(ws, widths)
    finalize_sheet(ws)


# ============================================================
# SHEET BUILDERS
# ============================================================


def build_model_specification_sheet(
    wb: Workbook,
    pages: list[DocumentationPage],
    summary_rows: list[tuple[str, str]],
    target_model: str,
    notes: list[tuple[str, str]],
) -> None:
    definition_page = find_best_model_definition_page(pages)
    model_definition = extract_definition(definition_page.content) or definition_page.summary
    update_frequency = extract_update_frequency(definition_page.content)

    column_specs_note = dict(notes)
    column_sources = [
        col.source_dependency
        for col in build_column_specs(pages)
        if col.source_dependency and col.source_dependency != "Not documented in CastorDoc export"
    ]

    golden_source = ", ".join(dedupe_keep_order(column_sources)) or "Not documented in CastorDoc export"

    rows = [
        ["Target model", target_model],
        ["Business purpose", model_definition],
        ["Suggested grain", column_specs_note.get("Suggested grain", "To be confirmed")],
        ["Suggested primary key", column_specs_note.get("Suggested primary key", "To be confirmed")],
        ["Refresh / update frequency", update_frequency or "Not documented in CastorDoc export"],
        ["Golden source", golden_source],
        ["Export folder", get_summary_value(summary_rows, "OUTPUT_DIR", "")],
        ["Saved documentation pages", get_summary_value(summary_rows, "SAVED_URLS", str(len(pages)))],
        ["Failed pages", get_summary_value(summary_rows, "FAILED_URLS", "0")],
        ["Definition evidence page", display_page_name(definition_page)],
        ["Definition evidence URL", definition_page.url],
    ]

    build_table_sheet(
        wb,
        "Model Specification",
        ["Specification Item", "Value"],
        rows,
        {"A": 35, "B": 120},
        "ModelSpecificationTable",
    )


def build_column_specification_sheet(wb: Workbook, specs: list[ColumnSpec]) -> None:
    headers = [
        "Column Name",
        "Business Name",
        "Suggested dbt Column Name",
        "Suggested Type",
        "Nullable?",
        "Source / Dependency",
        "Calculation Rule",
        "Business Definition",
        "Example Values",
        "Data Quality Rules",
        "Open Questions",
        "Evidence Page",
        "Evidence URL",
    ]

    rows = [
        [
            spec.column_name,
            spec.business_name,
            spec.suggested_dbt_name,
            spec.suggested_type,
            spec.nullable,
            spec.source_dependency,
            spec.calculation_rule,
            spec.business_definition,
            spec.example_values,
            spec.data_quality_rules,
            spec.open_questions,
            spec.evidence_page,
            spec.evidence_url,
        ]
        for spec in specs
    ]

    build_table_sheet(
        wb,
        "Column Specification",
        headers,
        rows,
        {
            "A": 34,
            "B": 34,
            "C": 34,
            "D": 20,
            "E": 16,
            "F": 35,
            "G": 55,
            "H": 70,
            "I": 45,
            "J": 55,
            "K": 55,
            "L": 40,
            "M": 90,
        },
        "ColumnSpecificationTable",
    )


def build_business_requirements_sheet(wb: Workbook, requirements: list[RequirementSpec]) -> None:
    headers = [
        "Requirement ID",
        "Requirement Type",
        "Requirement",
        "Expected Result",
        "Priority",
        "Evidence Page",
        "Evidence URL",
    ]

    rows = [
        [
            req.requirement_id,
            req.requirement_type,
            req.requirement,
            req.expected_result,
            req.priority,
            req.evidence_page,
            req.evidence_url,
        ]
        for req in requirements
    ]

    build_table_sheet(
        wb,
        "Business Requirements",
        headers,
        rows,
        {"A": 18, "B": 32, "C": 75, "D": 75, "E": 16, "F": 40, "G": 90},
        "BusinessRequirementsTable",
    )


def build_calculations_rules_sheet(wb: Workbook, rules: list[RuleSpec]) -> None:
    headers = [
        "Rule ID",
        "Field / Metric",
        "Rule Type",
        "Formula / Logic",
        "Dependencies",
        "Implementation Notes",
        "Evidence Page",
        "Evidence URL",
    ]

    rows = [
        [
            rule.rule_id,
            rule.field_or_metric,
            rule.rule_type,
            rule.formula_or_logic,
            rule.dependencies,
            rule.implementation_notes,
            rule.evidence_page,
            rule.evidence_url,
        ]
        for rule in rules
    ]

    build_table_sheet(
        wb,
        "Calculations & Rules",
        headers,
        rows,
        {"A": 16, "B": 35, "C": 26, "D": 70, "E": 40, "F": 55, "G": 40, "H": 90},
        "CalculationsRulesTable",
    )


def build_data_quality_tests_sheet(wb: Workbook, tests: list[TestSpec]) -> None:
    headers = [
        "Test Name",
        "Column(s)",
        "Test Type",
        "Severity",
        "Expected Result",
        "Suggested dbt Test",
        "Evidence Page",
        "Evidence URL",
    ]

    rows = [
        [
            test.test_name,
            test.columns,
            test.test_type,
            test.severity,
            test.expected_result,
            test.suggested_dbt_test,
            test.evidence_page,
            test.evidence_url,
        ]
        for test in tests
    ]

    build_table_sheet(
        wb,
        "Data Quality Tests",
        headers,
        rows,
        {"A": 45, "B": 45, "C": 28, "D": 16, "E": 70, "F": 70, "G": 40, "H": 90},
        "DataQualityTestsTable",
    )


def build_sources_dependencies_sheet(wb: Workbook, dependencies: list[DependencySpec]) -> None:
    headers = [
        "Target Model",
        "Dependency Type",
        "Dependency Name",
        "Dependency Role",
        "Join / Usage",
        "Evidence Page",
        "Evidence URL",
    ]

    rows = [
        [
            dep.target_model,
            dep.dependency_type,
            dep.dependency_name,
            dep.dependency_role,
            dep.join_or_usage,
            dep.evidence_page,
            dep.evidence_url,
        ]
        for dep in dependencies
    ]

    build_table_sheet(
        wb,
        "Source & Dependencies",
        headers,
        rows,
        {"A": 35, "B": 30, "C": 45, "D": 75, "E": 75, "F": 40, "G": 90},
        "SourcesDependenciesTable",
    )


def build_implementation_notes_sheet(wb: Workbook, notes: list[tuple[str, str]]) -> None:
    rows = [[topic, value] for topic, value in notes]

    build_table_sheet(
        wb,
        "Implementation Notes",
        ["Topic", "Notes"],
        rows,
        {"A": 35, "B": 120},
        "ImplementationNotesTable",
    )


def build_raw_evidence_sheet(wb: Workbook, pages: list[DocumentationPage]) -> None:
    headers = [
        "File",
        "Display Name",
        "Original Title",
        "Source Type",
        "Depth",
        "URL",
        "Parent URL",
        "ReadMe Content",
    ]

    rows = [
        [
            page.file_name,
            display_page_name(page),
            page.title,
            page.source_type,
            page.depth,
            page.url,
            page.parent_url,
            page.content,
        ]
        for page in pages
    ]

    build_table_sheet(
        wb,
        "Raw Evidence",
        headers,
        rows,
        {"A": 36, "B": 45, "C": 25, "D": 20, "E": 10, "F": 90, "G": 90, "H": 120},
        "RawEvidenceTable",
    )


def build_export_summary_sheet(wb: Workbook, summary_rows: list[tuple[str, str]], raw_summary_text: str) -> None:
    rows = [[key, value] for key, value in summary_rows]

    if raw_summary_text:
        rows.append(["RAW_EXPORT_SUMMARY", raw_summary_text])

    build_table_sheet(
        wb,
        "Export Summary",
        ["Key", "Value"],
        rows,
        {"A": 35, "B": 120},
        "ExportSummaryTable",
    )


def remove_default_sheet(wb: Workbook) -> None:
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])


# ============================================================
# PUBLIC API
# ============================================================


def build_excel_overview(export_folder: Path, output_file: Path | None = None) -> Path:
    if not export_folder.exists():
        raise FileNotFoundError(f"Export folder not found: {export_folder}")

    txt_paths = sorted(
        path
        for path in export_folder.glob("*.txt")
        if path.name != "export_summary.txt"
    )

    if not txt_paths:
        raise ValueError(f"No documentation .txt files found in: {export_folder}")

    pages = [parse_documentation_file(path) for path in txt_paths]

    summary_path = export_folder / "export_summary.txt"
    summary_rows = parse_summary_file(summary_path)
    raw_summary_text = summary_path.read_text(encoding="utf-8") if summary_path.exists() else ""

    target_model = infer_target_model_name(pages)
    column_specs = build_column_specs(pages)
    requirements = build_requirement_specs(pages)
    rules = build_rule_specs(column_specs)
    tests = build_test_specs(column_specs, requirements)
    dependencies = build_dependency_specs(pages, target_model, column_specs)
    notes = build_implementation_notes(target_model, column_specs, requirements, tests)

    wb = Workbook()

    build_model_specification_sheet(wb, pages, summary_rows, target_model, notes)
    build_column_specification_sheet(wb, column_specs)
    build_business_requirements_sheet(wb, requirements)
    build_calculations_rules_sheet(wb, rules)
    build_data_quality_tests_sheet(wb, tests)
    build_sources_dependencies_sheet(wb, dependencies)
    build_implementation_notes_sheet(wb, notes)
    build_raw_evidence_sheet(wb, pages)
    build_export_summary_sheet(wb, summary_rows, raw_summary_text)

    remove_default_sheet(wb)

    if output_file is None:
        output_file = export_folder / "castordoc_model_specification.xlsx"

    wb.save(output_file)
    return output_file


def main() -> None:
    args = parse_args()

    export_folder = Path(args.export_folder)
    output_file = Path(args.output_file) if args.output_file else None

    generated_file = build_excel_overview(
        export_folder=export_folder,
        output_file=output_file,
    )

    print(f"Excel generated: {generated_file}", flush=True)


if __name__ == "__main__":
    main()